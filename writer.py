# -*- coding: utf-8 -*-
"""
Writer — Xuất file DOCX đề trộn + Excel bảng đáp án.

Chạy debug:
    python writer.py input/de_goc.docx --test
"""

import os
import sys
import re
from copy import deepcopy
from typing import List

from docx import Document
from lxml import etree
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from models import ExamDocument, Question

# Đảm bảo encoding UTF-8
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8')

# Namespace
WML_NS = 'http://schemas.openxmlformats.org/wordprocessingml/2006/main'


# ============================================================
# DOCX Writer
# ============================================================

def _replace_number_in_text(xml_element, old_number: int, new_number: int):
    """
    Thay thế số câu hỏi trong XML element.
    Tìm pattern "Câu X" / "Question X" / "X." và thay thế số.
    """
    patterns = [
        (re.compile(rf'((?:Câu|Question)\s*){old_number}(\s*[.:\)])'), rf'\g<1>{new_number}\2'),
        (re.compile(rf'^{old_number}(\s*[.:\)])'), rf'{new_number}\1'),
    ]

    for t_elem in xml_element.iter():
        tag = t_elem.tag.split('}')[-1] if '}' in t_elem.tag else t_elem.tag
        if tag == 't' and t_elem.text:
            for pattern, replacement in patterns:
                new_text = pattern.sub(replacement, t_elem.text)
                if new_text != t_elem.text:
                    t_elem.text = new_text
                    return True  # Chỉ thay 1 lần
    return False


def _replace_option_letter_in_text(xml_element, old_letter: str, new_letter: str):
    """Thay thế chữ cái đáp án trong XML element."""
    if old_letter == new_letter:
        return

    for t_elem in xml_element.iter():
        tag = t_elem.tag.split('}')[-1] if '}' in t_elem.tag else t_elem.tag
        if tag == 't' and t_elem.text:
            # Thay "A." → "C." ở đầu text
            pattern = re.compile(rf'^{re.escape(old_letter)}(\s*[.)\s])')
            new_text = pattern.sub(rf'{new_letter}\1', t_elem.text)
            if new_text != t_elem.text:
                t_elem.text = new_text
                return


def _remove_correct_formatting(xml_element):
    """Xóa formatting đáp án đúng (tô đỏ, gạch chân) khỏi element."""
    for child in xml_element.iter():
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag == 'rPr':
            # Xóa color đỏ
            color_el = child.find(f'{{{WML_NS}}}color')
            if color_el is not None:
                color_val = color_el.get(f'{{{WML_NS}}}val', '') or color_el.get('val', '')
                if color_val and _is_red_hex(color_val):
                    child.remove(color_el)

            # Xóa underline
            u_el = child.find(f'{{{WML_NS}}}u')
            if u_el is not None:
                child.remove(u_el)


def _apply_correct_formatting(xml_element):
    """Thêm formatting đáp án đúng (gạch chân) vào element."""
    for child in xml_element.iter():
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag == 'rPr':
            # Thêm underline
            u_el = child.find(f'{{{WML_NS}}}u')
            if u_el is None:
                u_el = etree.SubElement(child, f'{{{WML_NS}}}u')
                u_el.set(f'{{{WML_NS}}}val', 'single')
            return

        if tag == 'r':
            # Nếu chưa có rPr, tạo mới
            rpr = child.find(f'{{{WML_NS}}}rPr')
            if rpr is None:
                rpr = etree.SubElement(child, f'{{{WML_NS}}}rPr')
                child.insert(0, rpr)  # rPr phải đứng đầu run
            u_el = etree.SubElement(rpr, f'{{{WML_NS}}}u')
            u_el.set(f'{{{WML_NS}}}val', 'single')
            return


def _is_red_hex(hex_color: str) -> bool:
    """Kiểm tra mã hex có phải đỏ không."""
    hex_color = hex_color.strip().lstrip('#')
    if len(hex_color) != 6:
        return False
    try:
        r = int(hex_color[0:2], 16)
        g = int(hex_color[2:4], 16)
        b = int(hex_color[4:6], 16)
        return r > 180 and g < 100 and b < 100
    except ValueError:
        return False


def write_shuffled_docx(shuffled_exam: ExamDocument, template_path: str,
                         output_path: str):
    """
    Xuất 1 file DOCX đề trộn.

    Strategy: Clone toàn bộ document gốc (giữ styles, page setup),
    xóa body, rồi clone lại paragraphs theo thứ tự mới.
    """
    # 1. Mở template gốc
    doc = Document(template_path)
    body = doc.element.body

    # 2. Xóa toàn bộ body content
    for child in list(body):
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag == 'sectPr':
            continue  # Giữ section properties (page size, margins)
        body.remove(child)

    # 3. Clone header paragraphs
    for elem in shuffled_exam.header_elements:
        body.append(deepcopy(elem))

    # 4. Clone từng câu hỏi theo thứ tự mới
    # Theo dõi element IDs đã ghi để tránh trùng lặp (inline options)
    written_elem_ids = set()

    for question in shuffled_exam.questions:
        # Clone context (đoạn văn đọc hiểu)
        for ctx_elem in question.context_elements:
            eid = id(ctx_elem)
            if eid not in written_elem_ids:
                body.append(deepcopy(ctx_elem))
                written_elem_ids.add(eid)

        if question.is_inline:
            # ═══ INLINE OPTIONS: tất cả A/B/C/D trên cùng paragraph ═══
            # Chỉ ghi 1 lần toàn bộ paragraph (stem + options)
            for stem_elem in question.stem_elements:
                eid = id(stem_elem)
                if eid not in written_elem_ids:
                    cloned = deepcopy(stem_elem)
                    # Thay số câu
                    if question.new_number and question.new_number != question.original_number:
                        _replace_number_in_text(cloned, question.original_number, question.new_number)
                    
                    # Nếu đáp án bị trộn → cần sắp xếp lại runs trong paragraph
                    if question.options_were_shuffled:
                        _reorder_inline_options(cloned, question)
                    
                    body.append(cloned)
                    written_elem_ids.add(eid)
        else:
            # ═══ DÒNG RIÊNG: stem + từng option trên dòng riêng ═══
            # Clone stem (thân câu hỏi)
            for stem_elem in question.stem_elements:
                cloned = deepcopy(stem_elem)
                # Thay số câu
                if question.new_number and question.new_number != question.original_number:
                    _replace_number_in_text(cloned, question.original_number, question.new_number)
                body.append(cloned)

            # Clone options theo thứ tự mới
            for letter in ['A', 'B', 'C', 'D']:
                opt = question.options.get(letter)
                if not opt:
                    continue

                for opt_elem in opt.elements:
                    cloned = deepcopy(opt_elem)

                    # Thay chữ cái đáp án nếu đã trộn
                    if question.options_were_shuffled:
                        _replace_option_letter_in_text(cloned, opt.original_letter, letter)

                        # Xử lý formatting đáp án đúng
                        if opt.is_correct:
                            _apply_correct_formatting(cloned)
                        else:
                            _remove_correct_formatting(cloned)
                    # Nếu không trộn đáp án → giữ nguyên formatting gốc

                    body.append(cloned)

    # 5. Save
    doc.save(output_path)
    print(f"  📄 Đã xuất: {os.path.basename(output_path)}")


def _reorder_inline_options(para_element, question: Question):
    """
    Sắp xếp lại runs trong paragraph chứa inline options khi đáp án bị trộn.
    
    Cấu trúc paragraph: 
      [stem runs] [tab] [A. ] [content_A_runs] [tab] [B. ] [content_B_runs] ...
    
    Khi trộn, cần đổi vị trí nội dung giữa các options.
    Strategy đơn giản: Tìm và đổi letter + color cho run chứa "X. "
    """
    if not question.option_mapping:
        return
    
    # Option mapping: new_letter → old_letter
    # VD: {'A': 'C', 'B': 'A', 'C': 'D', 'D': 'B'}
    # Nghĩa là: vị trí A mới chứa nội dung gốc C, v.v.
    
    option_letter_pattern = re.compile(r'^([A-D])\s*[.)]\s*$', re.IGNORECASE)
    
    # Bước 1: Tìm tất cả runs chứa option letters và đánh dấu
    letter_runs = {}  # {original_letter: run_element}
    for child in para_element:
        tag = child.tag.split('}')[-1] if '}' in child.tag else child.tag
        if tag != 'r':
            continue
        
        # Lấy text của run
        run_text = ''
        for t in child.iter():
            t_tag = t.tag.split('}')[-1] if '}' in t.tag else t.tag
            if t_tag == 't' and t.text:
                run_text += t.text
        
        m = option_letter_pattern.match(run_text.strip())
        if m:
            letter = m.group(1).upper()
            letter_runs[letter] = child
    
    # Bước 2: Cập nhật formatting cho letter runs
    # Đáp án đúng mới: letter nào có question.options[letter].is_correct = True
    for letter, run_elem in letter_runs.items():
        opt = question.options.get(letter)
        if not opt:
            continue
        
        rpr = run_elem.find(f'{{{WML_NS}}}rPr')
        if rpr is None:
            continue
        
        if opt.is_correct:
            # Đặt màu đỏ cho đáp án đúng mới
            color_el = rpr.find(f'{{{WML_NS}}}color')
            if color_el is None:
                color_el = etree.SubElement(rpr, f'{{{WML_NS}}}color')
            color_el.set(f'{{{WML_NS}}}val', 'FF0000')
            # Xóa themeColor nếu có
            for attr in list(color_el.attrib):
                if 'themeColor' in attr:
                    del color_el.attrib[attr]
        else:
            # Xóa màu đỏ nếu đáp án này không còn đúng
            color_el = rpr.find(f'{{{WML_NS}}}color')
            if color_el is not None:
                color_val = color_el.get(f'{{{WML_NS}}}val', '') or color_el.get('val', '')
                if color_val and _is_red_hex(color_val):
                    # Đặt về màu đen
                    color_el.set(f'{{{WML_NS}}}val', '000000')


# ============================================================
# Excel Answer Key Writer
# ============================================================

def write_answer_key_excel(original_exam: ExamDocument,
                           shuffled_exams: List[ExamDocument],
                           output_path: str):
    """
    Xuất bảng đáp án ra file Excel.

    Columns: Câu | Đề gốc | Mã 001 | Mã 002 | ...
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Đáp án"

    # Styles
    header_font = Font(bold=True, size=12)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=12, color="FFFFFF")
    center = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Header row
    headers = ["Câu", "Đề gốc"]
    for i in range(len(shuffled_exams)):
        headers.append(f"Mã {i + 1:03d}")

    for col, header_text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header_text)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    # Data rows
    max_questions = max(
        original_exam.total_questions,
        max((sh.total_questions for sh in shuffled_exams), default=0)
    )

    for q_num in range(1, max_questions + 1):
        row = q_num + 1

        # Câu số
        ws.cell(row=row, column=1, value=q_num).alignment = center
        ws.cell(row=row, column=1).border = thin_border

        # Đáp án gốc
        orig_q = original_exam.get_question(q_num)
        orig_ans = orig_q.correct_answer if orig_q else ""
        cell = ws.cell(row=row, column=2, value=orig_ans)
        cell.alignment = center
        cell.border = thin_border
        cell.font = Font(bold=True)

        # Đáp án từng đề trộn
        for sh_idx, sh_exam in enumerate(shuffled_exams):
            col = sh_idx + 3

            # Tìm câu hỏi có new_number == q_num trong đề trộn
            sh_ans = ""
            for sq in sh_exam.questions:
                if sq.new_number == q_num:
                    sh_ans = sq.correct_answer
                    break

            cell = ws.cell(row=row, column=col, value=sh_ans)
            cell.alignment = center
            cell.border = thin_border

    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 12
    for i in range(len(shuffled_exams)):
        col_letter = openpyxl.utils.get_column_letter(i + 3)
        ws.column_dimensions[col_letter].width = 12

    # Sheet 2: Chi tiết mapping câu
    ws2 = wb.create_sheet("Chi tiết mapping")
    ws2_headers = ["Đề", "Câu mới", "Câu gốc", "Đáp án gốc", "Đáp án mới", "Trộn đáp án"]
    for col, h in enumerate(ws2_headers, 1):
        cell = ws2.cell(row=1, column=col, value=h)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.alignment = center
        cell.border = thin_border

    row = 2
    for sh_idx, sh_exam in enumerate(shuffled_exams):
        for q in sh_exam.questions:
            ws2.cell(row=row, column=1, value=f"Mã {sh_idx + 1:03d}").border = thin_border
            ws2.cell(row=row, column=2, value=q.new_number).border = thin_border
            ws2.cell(row=row, column=3, value=q.original_number).border = thin_border
            ws2.cell(row=row, column=4, value=q.original_correct_answer).border = thin_border
            ws2.cell(row=row, column=5, value=q.correct_answer).border = thin_border
            ws2.cell(row=row, column=6, value="Có" if q.options_were_shuffled else "Không").border = thin_border
            row += 1

    wb.save(output_path)
    print(f"  📊 Đã xuất bảng đáp án: {os.path.basename(output_path)}")


# ============================================================
# Batch Writer
# ============================================================

def write_all_outputs(original_exam: ExamDocument,
                      shuffled_exams: List[ExamDocument],
                      template_path: str,
                      output_dir: str):
    """
    Xuất tất cả đề trộn + bảng đáp án.

    Args:
        original_exam: Đề gốc
        shuffled_exams: Danh sách đề trộn
        template_path: Path file DOCX gốc (dùng làm template)
        output_dir: Thư mục output
    """
    os.makedirs(output_dir, exist_ok=True)

    base_name = os.path.splitext(os.path.basename(template_path))[0]

    print(f"\n📁 Output: {output_dir}")

    # Xuất từng đề trộn
    for i, sh_exam in enumerate(shuffled_exams):
        output_path = os.path.join(output_dir, f"{base_name}_Ma{i + 1:03d}.docx")
        write_shuffled_docx(sh_exam, template_path, output_path)

    # Xuất bảng đáp án
    answer_path = os.path.join(output_dir, f"{base_name}_DapAn.xlsx")
    write_answer_key_excel(original_exam, shuffled_exams, answer_path)

    print(f"\n✅ Hoàn tất! Đã xuất {len(shuffled_exams)} đề + 1 file đáp án.")


# ============================================================
# Debug: chạy python writer.py <file.docx> --test
# ============================================================
if __name__ == "__main__":
    print("=" * 60)
    print("  DEBUG: writer.py — Xuất DOCX + Excel")
    print("=" * 60)

    if len(sys.argv) < 2:
        print("\nCách dùng: python writer.py <file.docx> --test")
        print("Ví dụ:     python writer.py input/de_goc.docx --test")

        # Demo: tạo Excel mẫu
        print("\n--- Demo tạo Excel đáp án mẫu ---")
        from models import OptionData

        exam = ExamDocument(filepath="Các quy tắc trộn đề.docx", total_questions=5)
        for i in range(1, 6):
            correct = chr(65 + (i % 4))  # A, B, C, D, A
            q = Question(original_number=i, correct_answer=correct,
                         original_correct_answer=correct,
                         stem_text=f"Question {i}")
            for letter in 'ABCD':
                q.options[letter] = OptionData(letter, letter,
                                                f"Opt {letter}",
                                                letter == correct)
            exam.questions.append(q)

        # Tạo "shuffled" giả
        from copy import deepcopy
        sh1 = deepcopy(exam)
        for q in sh1.questions:
            q.new_number = q.original_number
            q.correct_answer = chr(65 + ((ord(q.correct_answer) - 65 + 1) % 4))
            q.options_were_shuffled = True

        sh2 = deepcopy(exam)
        for q in sh2.questions:
            q.new_number = q.original_number
            q.correct_answer = chr(65 + ((ord(q.correct_answer) - 65 + 2) % 4))
            q.options_were_shuffled = True

        output_dir = os.path.join(os.path.dirname(__file__), "output")
        os.makedirs(output_dir, exist_ok=True)
        excel_path = os.path.join(output_dir, "demo_DapAn.xlsx")
        write_answer_key_excel(exam, [sh1, sh2], excel_path)

        print("\n✅ writer.py OK")
        sys.exit(0)

    # Test với file DOCX thật
    filepath = sys.argv[1]
    if not os.path.exists(filepath):
        print(f"❌ File không tồn tại: {filepath}")
        sys.exit(1)

    from parser import parse_docx
    from shuffler import generate_unique_exams
    from models import GroupConfig

    exam = parse_docx(filepath)

    # Trộn 2 đề test
    configs = [GroupConfig(group_type=3, question_ranges=[(1, exam.total_questions)])]
    shuffled = generate_unique_exams(exam, configs, 2)

    if shuffled:
        output_dir = os.path.join(os.path.dirname(filepath), "output")
        write_all_outputs(exam, shuffled, filepath, output_dir)
    else:
        print("❌ Không tạo được đề trộn!")

    print("\n✅ writer.py OK")
